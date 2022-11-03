import vk_api
from openpyxl import load_workbook
from vk_api.bot_longpoll import VkBotLongPoll
from vk_api.bot_longpoll import VkBotEventType
from vk_api.keyboard import VkKeyboard
from vk_api.keyboard import VkKeyboardColor
import requests
from my_token import my_token
from stepik_res import client_id_pr
from stepik_res import client_secret_pr
import json
import re
import os.path, time
import vk
from datetime import datetime
import collections
def keyboard_init():
  key_board = VkKeyboard(one_time = False, inline = False)
  key_board.add_button(
    label = "Узнать баллы на Stepik",
    color=VkKeyboardColor.SECONDARY,
    payload={'button': 1}
  )
  key_board.add_line()
  key_board.add_button(
    label = "Узнать текущее количество жизней",
    color=VkKeyboardColor.POSITIVE,
    payload={'button': 2}
  )
  key_board.add_button(
    label = "Дедлайн",
    color=VkKeyboardColor.NEGATIVE,
    payload={'button': 3}
  )
  key_board.add_line()
  key_board.add_button(
    label = "Письмо куратору в лс,если долго ждешь",
    color=VkKeyboardColor.SECONDARY,
    payload={'button': 4}
  )
  return key_board.get_keyboard()
#vk_api
group_id = 215643440
vk_session = vk_api.VkApi(token = my_token)
session_api = vk_session.get_api()
longpoll = VkBotLongPoll(vk_session, group_id)
def send_message(user_id, message):
  message_payload = {'user_id': user_id, 'random_id': 0, **message}
  vk_session.method('messages.send', message_payload)
def get_members(group_id):
  group_getMembers = {'group_id': group_id, 'offset': 0}
  return vk_session.method('groups.getMembers', group_getMembers)
all_members = get_members(group_id)['items']
ALL_MEMBERS_CURRENT_SESSION = []
#own vk_id
#screen_name - короткое имя после @
own_vk_id = vk_session.method('utils.resolveScreenName', {'screen_name': 'suyuki_cream'})['object_id']
print(own_vk_id)
#Excel info members
flash_wb = load_workbook('./a8.xlsx')
flash_gr = flash_wb.get_sheet_by_name('Лист1')
name_id_step = {} # 'vk_id': stepik_id
name_life_am = {} # 'name': amount_life
for i in range(1, get_members(group_id)['count'] + 1):
  tmp = flash_gr.cell(row = i, column = 3).value
  tmp1 = flash_gr.cell(row = i, column = 5).value
  if tmp is not None:
    if int(tmp[tmp.find("id")+2:])in all_members:
      vk_id = tmp[tmp.find("id")+2:]
      name_id_step[vk_id] = str(flash_gr.cell(row = i, column = 7).value)
      name_life_am[vk_id] = str(tmp1)
    else:
      print(flash_gr.cell(row = i, column = 2).value,
            "Такого ученика нет в твоей группе")
#Stepik info
client_id = client_id_pr
client_secret = client_secret_pr
auth = requests.auth.HTTPBasicAuth(client_id, client_secret)
response = requests.post('https://stepik.org/oauth2/token/',
                         data={'grant_type': 'client_credentials'},
                         auth=auth)
token = response.json().get('access_token', None)
if not token:
    print('Unable to authorize with provided credentials')
    exit(1)
class_api_call = 'https://stepik.org/api/classes/37459'
classes = requests.get(class_api_call, headers = {'Authorization': 'Bearer '+ token}).json()
'''Не нужно, потому что у степика апи...
students_api_call = 'https://stepik.org/api/students?klass=37459'
students = requests.get(students_api_call, headers = {'Authorization': 'Bearer '+ token}).json()
for i in range(0, classes['classes'][0]['students_count']):
  print(students['students'][i]['user'])

stepik_id_list = [0]*classes['classes'][0]['students_count']
name_stepik_id = {}
for i in range(0, 15): #15 поменять на classes['classes'][0]['students_count']
  stepik_id = students['students'][i]['user']
  find_stepik_name = 'https://stepik.org/users/' + str(stepik_id)
  stepik_profile = requests.get(find_stepik_name, headers = {'Authorization': 'Bearer '+ token})
  ''''''
  хз, почему не пашет регулярка
  stepik_name = re.findall(r'(<title>(.*)?<\/title>)', stepik_profile.text)
  ''''''
  stepik_name = [stepik_profile.text[k] for k in range(stepik_profile.text.find("<title>")+10,
                                                 stepik_profile.text.find("</title>")-10)]    
  stepik_name = ''.join(stepik_name)
  name_stepik_id[stepik_name] = stepik_id
'''
  #Excel stepik total
tmp_sit_stepik = load_workbook('./Текущая ситуация.xlsx')
tmp_sit = tmp_sit_stepik.get_sheet_by_name('Sheet1')
name_and_total = {}
for i in range(1, classes['classes'][0]['students_count'] + 2):
  tmp = tmp_sit.cell(row = i, column = 1).value
  if (str(tmp) in name_id_step.values()) == False and str(tmp) != "user_id":
    if str(tmp) != 'user_id':
      print('Этого человека нет в твоей группе в вк')
  else:
     for vk_id, step_id in name_id_step.items():
      if step_id == str(tmp):
          #При обновлении отчета степика поменять column = 158 на номер колонки total
          #Чтобы узнать номер колонки: Параметры->Формулы->Стиль рассылок R1C1
          #Если не поменялся - не менять
         name_id_step[vk_id] = str(round(tmp_sit.cell(row = i, column = 158).value,2))
name_id_step = dict(sorted(name_id_step.items(), key=lambda item: float(item[1])))
def total_pos_onstep(tmp):
   if tmp in name_id_step.keys() and tmp in total_sorted.keys():
    res = 'Сумма твоих баллов на степике: ' + name_id_step[tmp] + ". По группе ты на " + str(classes['classes'][0]['students_count']  - list(name_id_step.keys()).index(tmp))+ " месте."
    return res
   else:
     print('Что-то пошло не так')
   total_sorted = dict(sorted(name_id_step.items(), key=lambda item: item[1]))
   last_modified = time.ctime(os.path.getmtime('./Текущая ситуация.xlsx'))
  #Файл с дедлайнами
def find_next_dd():
   f = open('ddays.txt', encoding='utf-8')
   cur_time = datetime.now()
   diff_time = {}
   info_dday_time = {}
   for line in f:
     pr = []
     for i in range(0, len(line)):
       if line[i] == ' ':
         pr.append(i)
     year = line[:pr[0]]
     month = line[pr[0] + 1: pr[1]]
     day = line[pr[1] + 1: pr[2]]
     info = line[pr[2] + 1:]
     dday_time = datetime(int(year), int(month), int(day))
     diff_time[dday_time] = cur_time - dday_time
     info_dday_time[dday_time] = info
   diff_time = sorted(diff_time.items())
   info_dday_time = sorted(info_dday_time.items())
   info_dday = info_dday_time[0]
   info_dday = info_dday[1][:len(info_dday[1])-1]
   needed_item_time = diff_time[0]
   delta_needed_item_time = needed_item_time[1]
   f.close()
   return [delta_needed_item_time, info_dday]
  #return str(delta_needed_item_time)[1:])
def total_life_course(tmp):
  if tmp in name_life_am.keys():
    print(name_life_am.get(tmp))
    return name_life_am.get(tmp)
  #Вопрос-ответ
def button_click(user_id, payload):
  res = json.loads(payload)
  if 'button' in res:
    button_index = res['button']
    if button_index == 1:
      send_message(user_id, {'message' : total_pos_onstep(str(user_id)) +
                               " Last stepik_places update: " + last_modified})
    if button_index == 2:
      send_message(user_id, {'message' : "Текущее количество жизней: " + 
                               total_life_course(str(user_id))})
    if button_index == 3:
      send_message(user_id, {'message' : "Следующий дедлайн через: " + str(find_next_dd()[0])[1:] + " по теме: " +
                               find_next_dd()[1]})
    if button_index == 4:
      send_message(own_vk_id, {'message' : "Гена, тебе написали в группе"})
for event in longpoll.listen():
  if event.type == VkBotEventType.MESSAGE_NEW:
    user_message = event.object.message['text'].lower()
    user_id = event.object.message['from_id']
    print(ALL_MEMBERS_CURRENT_SESSION)
    if user_id not in ALL_MEMBERS_CURRENT_SESSION:
      ALL_MEMBERS_CURRENT_SESSION.append(user_id)
      send_message(user_id, {'message': 'Flash_group_bot', 'keyboard': keyboard_init()})
      print(user_message + " from " + str(user_id))
    if 'payload' in event.object.message:
      button_click(user_id, event.object.message['payload'])

      
#НЕ ЗАБЫТЬ ПРО response.status и тд                               
#А что с валерой...
